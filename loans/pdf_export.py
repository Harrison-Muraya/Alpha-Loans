"""
loans/pdf_export.py  —  WeasyPrint + openpyxl exports

Install:
    pip install weasyprint openpyxl

WeasyPrint on Ubuntu/Debian also needs system libs:
    sudo apt install libpango-1.0-0 libpangoft2-1.0-0 libcairo2

WeasyPrint on Windows:
    Follow https://doc.courtbouillon.org/weasyprint/stable/first_steps.html#windows
"""
from io import BytesIO
from django.template.loader import render_to_string
from django.utils import timezone


# ── PDF via WeasyPrint ────────────────────────────────────────────────────────

def generate_borrower_statement(borrower, loans) -> bytes:
    """
    Render the loans/pdf_statement.html template and convert to PDF.
    Returns PDF bytes.
    """
    try:
        from weasyprint import HTML, CSS
    except ImportError:
        raise ImportError(
            "WeasyPrint is not installed. Run: pip install weasyprint\n"
            "On Linux also run: sudo apt install libpango-1.0-0 libpangoft2-1.0-0 libcairo2"
        )

    # Build context
    loans_list = list(loans)
    total_loaned   = sum(l.amount           for l in loans_list)
    total_due      = sum(l.total_amount_due() for l in loans_list)
    total_paid_amt = sum(l.total_paid()      for l in loans_list)
    total_balance  = sum(l.balance_remaining() for l in loans_list)

    context = {
        'borrower':      borrower,
        'loans':         loans_list,
        'total_loaned':  total_loaned,
        'total_due':     total_due,
        'total_paid':    total_paid_amt,
        'total_balance': total_balance,
        'generated_at':  timezone.now(),
    }

    html_string = render_to_string('loans/pdf_statement.html', context)
    pdf_bytes   = HTML(string=html_string).write_pdf()
    return pdf_bytes


# ── Excel via openpyxl ────────────────────────────────────────────────────────

def generate_loans_excel(loans) -> bytes:
    """
    Generate an Excel workbook with Loans + Payments sheets.
    Returns .xlsx bytes.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()

    BLUE_FILL  = PatternFill('solid', fgColor='2D5BE3')
    ALT_FILL   = PatternFill('solid', fgColor='F0EDE8')
    HDR_FONT   = Font(color='FFFFFF', bold=True, size=10)
    BOLD_FONT  = Font(bold=True)
    CENTER     = Alignment(horizontal='center')

    def style_header_row(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = HDR_FONT
            cell.fill      = BLUE_FILL
            cell.alignment = CENTER
        ws.freeze_panes = 'A2'

    def auto_width(ws, count):
        for col in range(1, count + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet 1: Loans ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Loans'
    hdrs1 = [
        'Loan ID', 'Borrower', 'Username', 'Amount (KES)',
        'Interest %', 'Penalty %', 'Total Due (KES)',
        'Total Paid (KES)', 'Balance (KES)',
        'Status', 'Issue Date', 'Due Date', 'Paid?',
    ]
    style_header_row(ws1, hdrs1)

    for r, loan in enumerate(loans, 2):
        fill = ALT_FILL if r % 2 == 0 else PatternFill()
        row  = [
            loan.id,
            loan.borrower.get_full_name() or loan.borrower.username,
            loan.borrower.username,
            float(loan.amount),
            float(loan.interest_rate),
            float(loan.penalty_rate),
            float(loan.total_amount_due()),
            float(loan.total_paid()),
            float(loan.balance_remaining()),
            loan.get_status_display(),
            loan.issue_date.strftime('%Y-%m-%d'),
            loan.due_date.strftime('%Y-%m-%d'),
            'Yes' if loan.is_paid else 'No',
        ]
        for col, val in enumerate(row, 1):
            c = ws1.cell(row=r, column=col, value=val)
            c.fill = fill

    auto_width(ws1, len(hdrs1))

    # ── Sheet 2: Payments ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('Payments')
    hdrs2 = [
        'Loan ID', 'Borrower', 'Amount (KES)',
        'Method', 'Reference', 'Note', 'Paid At', 'Recorded By',
    ]
    style_header_row(ws2, hdrs2)

    r = 2
    for loan in loans:
        for p in loan.payments.all():
            fill = ALT_FILL if r % 2 == 0 else PatternFill()
            row  = [
                loan.id,
                loan.borrower.get_full_name() or loan.borrower.username,
                float(p.amount),
                p.get_method_display(),
                p.reference or '',
                p.note or '',
                p.paid_at.strftime('%Y-%m-%d %H:%M'),
                p.recorded_by.username if p.recorded_by else '',
            ]
            for col, val in enumerate(row, 1):
                c = ws2.cell(row=r, column=col, value=val)
                c.fill = fill
            r += 1

    auto_width(ws2, len(hdrs2))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()